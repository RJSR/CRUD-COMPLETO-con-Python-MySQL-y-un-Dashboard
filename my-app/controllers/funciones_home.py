
# Para subir archivo tipo foto al servidor
from werkzeug.utils import secure_filename
import uuid  # Modulo de python para crear un string

from conexion.conexionBD import connectionBD  # Conexión a BD

import datetime
import re
import os

from os import remove  # Modulo  para remover archivo
from os import path  # Modulo para obtener la ruta o directorio


import openpyxl  # Para generar el excel
# biblioteca o modulo send_file para forzar la descarga
from flask import send_file


def procesar_form_producto(dataForm):
    # Formateando Salario
    precio_sin_puntos = re.sub('[^0-9]+', '', dataForm['precio_dolar'])
    precio_sin_puntos_bsd = re.sub('[^0-9]+', '', dataForm['precio_bsd'])
    # convertir salario a INT
    precio_dolar_entero = int(precio_sin_puntos)
    precio_bsd_entero = int(precio_sin_puntos_bsd)

    
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:

                sql = "INSERT INTO tbl_productos (nombre_producto, marca_producto, cantidad, precio_dolar, precio_bsd) VALUES (%s, %s, %s, %s, %s)"

                # Creando una tupla con los valores del INSERT
                valores = (dataForm['nombre_producto'], dataForm['marca_producto'], dataForm['cantidad'],precio_dolar_entero, precio_bsd_entero)
                cursor.execute(sql, valores)

                conexion_MySQLdb.commit()
                resultado_insert = cursor.rowcount
                return resultado_insert

    except Exception as e:
        return f'Se produjo un error en procesar_form_producto: {str(e)}'


# def procesar_imagen_perfil(foto):
#     try:
#         # Nombre original del archivo
#         filename = secure_filename(foto.filename)
#         extension = os.path.splitext(filename)[1]

#         # Creando un string de 50 caracteres
#         nuevoNameFile = (uuid.uuid4().hex + uuid.uuid4().hex)[:100]
#         nombreFile = nuevoNameFile + extension

#         # Construir la ruta completa de subida del archivo
#         basepath = os.path.abspath(os.path.dirname(__file__))
#         upload_dir = os.path.join(basepath, f'../static/fotos_productos/')

#         # Validar si existe la ruta y crearla si no existe
#         if not os.path.exists(upload_dir):
#             os.makedirs(upload_dir)
#             # Dando permiso a la carpeta
#             os.chmod(upload_dir, 0o755)

#         # Construir la ruta completa de subida del archivo
#         upload_path = os.path.join(upload_dir, nombreFile)
#         foto.save(upload_path)

#         return nombreFile

#     except Exception as e:
#         print("Error al procesar archivo:", e)
#         return []


# Lista de productos
def sql_lista_productosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = (f"""
                    SELECT 
                        e.id_producto,
                        e.nombre_producto, 
                        e.marca_producto,
                        e.cantidad,
                        e.precio_dolar,
                        e.precio_bsd
                    FROM tbl_productos AS e
                    ORDER BY e.id_producto ASC
                    """)
                cursor.execute(querySQL,)
                productosBD = cursor.fetchall()
        return productosBD
    except Exception as e:
        print(
            f"Error en la función sql_lista_productosBD: {e}")
        return None


# Detalles del producto
def sql_detalles_productosBD(idproducto):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_producto,
                        e.nombre_producto, 
                        e.marca_producto,
                        e.cantidad,
                        e.precio_dolar,
                        e.precio_bsd,
                        DATE_FORMAT(e.fecha_registro, '%Y-%m-%d %h:%i %p') AS fecha_registro
                    FROM tbl_productos AS e
                    WHERE id_producto =%s
                    ORDER BY e.id_producto ASC
                    """)
                cursor.execute(querySQL, (idproducto))
                productosBD = cursor.fetchone()
        return productosBD
    except Exception as e:
        print(
            f"Error en la función sql_detalles_productosBD: {e}")
        return None


# Funcion productos Informe (Reporte)
def productosReporte():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = ("""
                    SELECT 
                        e.id_producto,
                        e.nombre_producto, 
                        e.marca_producto,
                        e.cantidad,
                        e.precio_dolar,
                        e.precio_bsd
                    FROM tbl_productos AS e
                    ORDER BY e.id_producto ASC
                    """)
                cursor.execute(querySQL,)
                productosBD = cursor.fetchall()
        return productosBD
    except Exception as e:
        print(
            f"Error en la función productosReporte: {e}")
        return None


def generarReporteExcel():
    dataproductos = productosReporte()
    wb = openpyxl.Workbook()
    hoja = wb.active

    # Agregar la fila de encabezado con los títulos
    cabeceraExcel = ("Nombre", "Marca", "Cantidad",
                     "Precio $", "Precio Bs")

    hoja.append(cabeceraExcel)

    # Formato para números en moneda colombiana y sin decimales
    formato_moneda_colombiana = '#,##0'

    # Agregar los registros a la hoja
    for registro in dataproductos:
        nombre_producto = registro['nombre_producto']
        marca_producto = registro['marca_producto']
        cantidad = registro['cantidad']
        precio_dolar = registro['precio_dolar']
        precio_bsd = registro['precio_bsd']


        # Agregar los valores a la hoja
        hoja.append((nombre_producto, marca_producto, cantidad, precio_dolar, precio_bsd))

        # Itera a través de las filas y aplica el formato a la columna G
        for fila_num in range(2, hoja.max_row + 1):
            columna = 4  # Columna G
            celda = hoja.cell(row=fila_num, column=columna)
            celda.number_format = formato_moneda_colombiana

    fecha_actual = datetime.datetime.now()
    archivoExcel = f"Reporte_productos_{fecha_actual.strftime('%Y_%m_%d')}.xlsx"
    carpeta_descarga = "../static/downloads-excel"
    ruta_descarga = os.path.join(os.path.dirname(
        os.path.abspath(__file__)), carpeta_descarga)

    if not os.path.exists(ruta_descarga):
        os.makedirs(ruta_descarga)
        # Dando permisos a la carpeta
        os.chmod(ruta_descarga, 0o755)

    ruta_archivo = os.path.join(ruta_descarga, archivoExcel)
    wb.save(ruta_archivo)

    # Enviar el archivo como respuesta HTTP
    return send_file(ruta_archivo, as_attachment=True)


def buscarproductoBD(search):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_producto,
                            e.nombre_producto, 
                            e.marca_producto,
                            e.cantidad,
                            e.precio_dolar,
                            e.precio_bsd
                        FROM tbl_productos AS e
                        WHERE e.nombre_producto LIKE %s 
                        ORDER BY e.id_producto ASC
                    """)
                search_pattern = f"%{search}%"  # Agregar "%" alrededor del término de búsqueda
                mycursor.execute(querySQL, (search_pattern,))
                resultado_busqueda = mycursor.fetchall()
                return resultado_busqueda

    except Exception as e:
        print(f"Ocurrió un error en def buscarproductoBD: {e}")
        return []


def buscarproductoUnico(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as mycursor:
                querySQL = ("""
                        SELECT 
                            e.id_producto,
                            e.nombre_producto, 
                            e.marca_producto,
                            e.cantidad,
                            e.precio_dolar,
                            e.precio_bsd
                        FROM tbl_productos AS e
                        WHERE e.id_producto =%s LIMIT 1
                    """)
                mycursor.execute(querySQL, (id,))
                producto = mycursor.fetchone()
                return producto

    except Exception as e:
        print(f"Ocurrió un error en def buscarproductoUnico: {e}")
        return []


def procesar_actualizacion_form(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                nombre_producto = data.form['nombre_producto']
                marca_producto = data.form['marca_producto']
                cantidad = data.form['cantidad']

                precio_sin_puntos_dolar = re.sub(
                    '[^0-9]+', '', data.form['precio_dolar'])
                precio_dolar = int(precio_sin_puntos_dolar)

                precio_sin_puntos_bsd = re.sub(
                    '[^0-9]+', '', data.form['precio_bsd'])
                precio_bsd = int(precio_sin_puntos_bsd)

                id_producto = data.form['id_producto']
                
                querySQL = """
                    UPDATE tbl_productos
                    SET 
                        nombre_producto = %s,
                        marca_producto = %s,
                        cantidad = %s,
                        precio_dolar = %s,
                        precio_bsd = %s
                    WHERE id_producto = %s
                """
                values = (nombre_producto, marca_producto, cantidad, precio_dolar, precio_bsd, id_producto)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None

def procesar_actualizacion_form_precio(data):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                
                querySQL = ("""
                    SELECT *
                        FROM tbl_productos
                        WHERE precio_dolar;
                    """)
                cursor.execute(querySQL)
                precio_dolar = cursor.fetchall()

                tasa = int(data.form['tasa'])

                precio_bsd = tasa * precio_dolar

                
                querySQL = """
                    UPDATE tbl_productos
                    SET 
                        precio_bsd = %s
                """
                values = (precio_bsd)

                cursor.execute(querySQL, values)
                conexion_MySQLdb.commit()

        return cursor.rowcount or []
    except Exception as e:
        print(f"Ocurrió un error en procesar_actualizacion_form: {e}")
        return None

# Lista de Usuarios creados
def lista_usuariosBD():
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "SELECT id, name_surname, email_user, created_user FROM users"
                cursor.execute(querySQL,)
                usuariosBD = cursor.fetchall()
        return usuariosBD
    except Exception as e:
        print(f"Error en lista_usuariosBD : {e}")
        return []


# Eliminar un producto
def eliminarproducto(id_producto):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM tbl_productos WHERE id_producto=%s"
                cursor.execute(querySQL, (id_producto,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount


        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarproducto : {e}")
        return []


# Eliminar usuario
def eliminarUsuario(id):
    try:
        with connectionBD() as conexion_MySQLdb:
            with conexion_MySQLdb.cursor(dictionary=True) as cursor:
                querySQL = "DELETE FROM users WHERE id=%s"
                cursor.execute(querySQL, (id,))
                conexion_MySQLdb.commit()
                resultado_eliminar = cursor.rowcount

        return resultado_eliminar
    except Exception as e:
        print(f"Error en eliminarUsuario : {e}")
        return []
